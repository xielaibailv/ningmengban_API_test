from openpyxl import load_workbook


# 创建用例对象
class Cases:
    def __init__(self):
        self.id = None
        self.title = None
        self.expected = None
        self.data = None


class DoExcel:
    """从Excel文件里读取数据"""

    def __init__(self, filename, sheet_name):
        self.filename = filename  # 工作簿
        self.sheet_name = sheet_name  # sheet名

    def open_sheet(self):
        excel = load_workbook(self.filename)  # 打开Excel
        sheet = excel[self.sheet_name]  # 打开sheet
        return excel, sheet

    # 读取数据
    def read_data(self):
        sheet = self.open_sheet()[1]
        # 把数据存到列表，每条数据是一个用例
        cases = []
        for i in range(2, sheet.max_row + 1):
            dict_data = {}  # 将请求用到的数据用字典格式存储
            case = Cases()
            if self.sheet_name == 'register':  # 第一个sheet，注册用例
                dict_data['mobilephone'] = sheet.cell(i, 3).value
                dict_data['regname'] = sheet.cell(i, 4).value
                dict_data['pwd'] = sheet.cell(i, 5).value

                case.id = sheet.cell(i, 1).value
                case.title = sheet.cell(i, 2).value
                case.expected = sheet.cell(i, 6).value

            elif self.sheet_name == 'login':  # 第2个sheet，登录用例
                dict_data['mobilephone'] = sheet.cell(i, 3).value
                dict_data['pwd'] = sheet.cell(i, 4).value

                case.id = sheet.cell(i, 1).value
                case.title = sheet.cell(i, 2).value
                case.expected = sheet.cell(i, 5).value

            elif self.sheet_name == 'recharge':  # 第3个sheet，充值用例
                dict_data['mobilephone'] = sheet.cell(i, 3).value
                dict_data['amount'] = sheet.cell(i, 4).value

                case.id = sheet.cell(i, 1).value
                case.title = sheet.cell(i, 2).value
                case.expected = sheet.cell(i, 5).value

            case.data = dict_data
            cases.append(case)
        return cases

    # 回写数据
    def write_data(self, row, column, value):
        wb = self.open_sheet()
        excel = wb[0]
        sheet = wb[1]
        sheet.cell(row, column).value = value
        excel.save(self.filename)


if __name__ == '__main__':
    wb = DoExcel('test_data.xlsx','register')
    cases = wb.read_data()
    for i in cases:
        print(i.data)



